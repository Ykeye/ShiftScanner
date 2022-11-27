import openpyxl, sys, time, random
from datetime import datetime, timedelta

typing_speed = 70 #wpm
def slow_type(t):
    for l in t:
        sys.stdout.write(l)
        sys.stdout.flush()
        time.sleep(random.random()*10.0/typing_speed)
    print('')

slow_type("And now, tell me where do you hide your corporate secret schedule: ")

for line in sys.stdin:

    file_location = line.strip()
    if 'Exit' == line.rstrip():
        break
    if line.rstrip():
        print(file_location + ", you say, ok let's have a look")
        break

wb = openpyxl.load_workbook(file_location)
sheet = wb['Лист1']
ws = wb.active

def get_shift_dates(surname):

    for row in tuple(ws.rows):
        row = str(row[1])
        row = row.split(".")
        row = row[1]
        row = row[:-1]
        cellvalue = sheet[f'{row}'].value
        rownumber = row[-2:]
        if cellvalue == str(surname):
            userrow = rownumber


    rowcount = ws.calculate_dimension()

    rowcount = (rowcount.split(":")[1])
    rowcount = rowcount[:2]


    rowcount =openpyxl.utils.cell.column_index_from_string(rowcount)



    resulting_dict = {}
    n_list = []
    d_list = []
    d2_list= []
    for col in range (3, int(rowcount)):
            result = ws.cell(row = int(userrow), column= int(col)).value

            if result == 'Н':

                    output_date =(ws.cell(row = 1, column = int((col)) -1).value)
                    output_date = output_date + timedelta(hours=20)
                    n_list+= [output_date]

            elif result == "Д":
                    output_date = (ws.cell(row=1, column=int(col)).value)
                    output_date = output_date + timedelta(hours=8)
                    d_list += [output_date]

            elif result == "Д2":
                   output_date=(ws.cell(row=1, column=int(col)).value)
                   output_date = output_date + timedelta(hours=8)
                   d2_list += [output_date]

    resulting_dict.update({'Н':n_list,'Д':d_list,'Д2':d2_list })


    return resulting_dict




