import openpyxl
from datetime import datetime, timedelta

wb = openpyxl.load_workbook("./shifts.xlsx")
sheet = wb['Лист1']
ws = wb.active
print(sheet['C1'].value)



for row in tuple(ws.rows):
    row = str(row[1])
    row = row.split(".")
    row = row[1]
    row = row[:-1]
    cellvalue = sheet[f'{row}'].value
    rownumber = row[-2:]
    if cellvalue == "Конев":
        userrow = rownumber
        print(cellvalue)
        print(rownumber)

rowcount = ws.calculate_dimension()
print(rowcount)
rowcount = (rowcount.split(":")[1])
rowcount = rowcount[:2]
print(rowcount)

rowcount =openpyxl.utils.cell.column_index_from_string(rowcount)
print(rowcount)
print(userrow)



for col in range (3, int(rowcount)):
   result = ws.cell(row = int(userrow), column= int(col)).value

   if result == 'Д' or result == 'Д2' or result == 'Н':
       print(result)
       if result == 'Н':
        output_date =(ws.cell(row = 1, column = int((col)) -1).value)

        output_date = output_date + timedelta(hours=20)
        print(output_date)

       else:
           output_date=(ws.cell(row=1, column=int(col)).value)
           output_date = output_date + timedelta(hours=8)
           print(output_date)






